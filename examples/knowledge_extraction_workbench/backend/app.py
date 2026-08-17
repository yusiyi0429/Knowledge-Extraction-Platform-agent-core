"""aiohttp API and static frontend host for the knowledge extraction workbench."""

from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import json
import os
import shutil
import zipfile
from datetime import datetime
from pathlib import Path, PurePath
from typing import Any
from urllib.parse import urlparse

import aiofiles
from aiohttp import web
from openpyxl import Workbook, load_workbook
from sqlmodel import func, select

from .config import SecretBox, Settings
from .errors import WorkbenchError, error_middleware
from .models import (
    AbilityProfile,
    AbilityMount,
    Asset,
    EvaluationRun,
    Exploration,
    ExplorationCandidate,
    ExtractionRound,
    FeedbackTask,
    Job,
    KnowledgeDocument,
    Material,
    ModelConnection,
    Revision,
    Scene,
    SkillVersion,
    Suggestion,
    new_id,
    utc_now,
)
from .model_runtime import OpenJiuwenKnowledgeModel, build_model_client_config, model_connection_error
from .pipeline import SUPPORTED_EXTENSIONS, parse_material, validate_skill_zip
from .service import REQUIRED_ASSET_KINDS, TERMINAL_JOB_STATUSES, WorkbenchService
from .store import ABILITY_SPECS, Store

APP_SETTINGS = web.AppKey("settings", Settings)
APP_STORE = web.AppKey("store", Store)
APP_SERVICE = web.AppKey("service", WorkbenchService)
APP_SECRET_BOX = web.AppKey("secret_box", SecretBox)


def _iso(value: datetime | None) -> str | None:
    return value.isoformat(timespec="seconds") + "Z" if value else None


def _json_loads(value: str, fallback: Any) -> Any:
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return fallback


async def _json_body(request: web.Request) -> dict[str, Any]:
    try:
        payload = await request.json()
    except (json.JSONDecodeError, web.HTTPBadRequest) as exc:
        raise WorkbenchError("REQUEST_JSON_INVALID", "请求 JSON 无效。", status=400) from exc
    if not isinstance(payload, dict):
        raise WorkbenchError("REQUEST_JSON_INVALID", "请求内容必须是 JSON 对象。", status=400)
    return payload


def _job_dict(job: Job) -> dict[str, Any]:
    return {
        "id": job.id,
        "kind": job.kind,
        "status": job.status,
        "phase": job.phase,
        "progress": job.progress,
        "seq": job.seq,
        "message": job.message,
        "scene_id": job.scene_id,
        "round_id": job.round_id,
        "exploration_id": job.exploration_id,
        "error": {
            "code": job.error_code,
            "message": job.error_message,
            "retryable": job.retryable,
        }
        if job.error_code
        else None,
        "created_at": _iso(job.created_at),
        "updated_at": _iso(job.updated_at),
    }


def _round_dict(round_row: ExtractionRound) -> dict[str, Any]:
    return {
        "id": round_row.id,
        "scene_id": round_row.scene_id,
        "version": round_row.version,
        "status": round_row.status,
        "subscenes": _json_loads(round_row.subscenes_json, []),
        "published_at": _iso(round_row.published_at),
        "created_at": _iso(round_row.created_at),
        "updated_at": _iso(round_row.updated_at),
    }


def _material_dict(material: Material) -> dict[str, Any]:
    return {
        "id": material.id,
        "round_id": material.round_id,
        "exploration_id": material.exploration_id,
        "name": material.name,
        "role": material.role,
        "extension": material.extension,
        "size_bytes": material.size_bytes,
        "sha256": material.sha256,
        "enabled": material.enabled,
        "created_at": _iso(material.created_at),
    }


def _document_dict(document: KnowledgeDocument) -> dict[str, Any]:
    return {
        "id": document.id,
        "round_id": document.round_id,
        "markdown": document.markdown,
        "structured": _json_loads(document.structured_json, {}),
        "revision": document.revision,
        "updated_at": _iso(document.updated_at),
    }


def _suggestion_dict(suggestion: Suggestion) -> dict[str, Any]:
    return {
        "id": suggestion.id,
        "round_id": suggestion.round_id,
        "base_revision": suggestion.base_revision,
        "old_text": suggestion.old_text,
        "new_text": suggestion.new_text,
        "explanation": suggestion.explanation,
        "source_refs": _json_loads(suggestion.source_refs_json, []),
        "status": suggestion.status,
        "created_at": _iso(suggestion.created_at),
        "resolved_at": _iso(suggestion.resolved_at),
    }


def _asset_dict(asset: Asset, current_revision: int | None = None) -> dict[str, Any]:
    path = Path(asset.file_path)
    return {
        "id": asset.id,
        "round_id": asset.round_id,
        "kind": asset.kind,
        "filename": asset.filename,
        "mime_type": asset.mime_type,
        "version": asset.version,
        "source_revision": asset.source_revision,
        "stale": current_revision is not None and asset.source_revision != current_revision,
        "synthetic": asset.synthetic,
        "size_bytes": path.stat().st_size if path.is_file() else 0,
        "created_at": _iso(asset.created_at),
        "download_url": f"/api/v1/assets/{asset.id}/download",
        "preview_url": f"/api/v1/assets/{asset.id}/preview",
    }


def _preview_cell(value: Any) -> str | int | float | bool | None:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    return str(value)


def _asset_base_preview(asset: Asset, mode: str) -> dict[str, Any]:
    return {
        "id": asset.id,
        "kind": asset.kind,
        "filename": asset.filename,
        "mode": mode,
        "download_url": f"/api/v1/assets/{asset.id}/download",
    }


def _model_dict(model: ModelConnection) -> dict[str, Any]:
    return {
        "id": model.id,
        "name": model.name,
        "provider": model.provider,
        "api_base": model.api_base,
        "model_name": model.model_name,
        "enabled": model.enabled,
        "has_api_key": bool(model.encrypted_api_key),
        "created_at": _iso(model.created_at),
        "updated_at": _iso(model.updated_at),
    }


def _skill_dict(skill: SkillVersion) -> dict[str, Any]:
    manifest = _json_loads(skill.manifest_json, {})
    return {
        "id": skill.id,
        "name": skill.name,
        "description": skill.description,
        "version": skill.version,
        "status": skill.status,
        "built_in": bool(manifest.get("built_in")),
        "kind": str(manifest.get("kind", "INSTANCE")),
        "read_only": bool(manifest.get("read_only")),
        "lineage_id": str(manifest.get("lineage_id", skill.id)),
        "source_skill_id": manifest.get("source_skill_id"),
        "source_name": str(manifest.get("source_name", "")),
        "scene_name": str(manifest.get("scene_name", "")),
        "notes": str(manifest.get("notes", "")),
        "manifest": manifest,
        "created_at": _iso(skill.created_at),
        "download_url": f"/api/v1/skills/{skill.id}/download",
    }


def _evaluation_dict(evaluation: EvaluationRun) -> dict[str, Any]:
    return {
        "id": evaluation.id,
        "round_id": evaluation.round_id,
        "model_connection_id": evaluation.model_connection_id,
        "job_id": evaluation.job_id,
        "dataset_name": evaluation.dataset_name,
        "dataset_kind": evaluation.dataset_kind,
        "status": evaluation.status,
        "sample_count": evaluation.sample_count,
        "correct_count": evaluation.correct_count,
        "wrong_count": max(0, evaluation.sample_count - evaluation.correct_count),
        "review_count": evaluation.review_count,
        "accuracy": evaluation.accuracy,
        "results": _json_loads(evaluation.results_json, []),
        "created_at": _iso(evaluation.created_at),
        "completed_at": _iso(evaluation.completed_at),
    }


def _feedback_task_dict(task: FeedbackTask) -> dict[str, Any]:
    cases = _json_loads(task.cases_json, [])
    return {
        "id": task.id,
        "round_id": task.round_id,
        "model_connection_id": task.model_connection_id,
        "job_id": task.job_id,
        "name": task.name,
        "task_type": task.task_type,
        "status": task.status,
        "source_filename": task.source_filename,
        "case_count": len(cases) if isinstance(cases, list) else 0,
        "cases": cases if isinstance(cases, list) else [],
        "promoted_round_id": task.promoted_round_id,
        "created_at": _iso(task.created_at),
        "updated_at": _iso(task.updated_at),
    }


async def health(_: web.Request) -> web.Response:
    return web.json_response({"status": "ok", "mode": "local-demo", "model_runtime": "configured-provider"})


async def api_not_found(_: web.Request) -> web.Response:
    raise WorkbenchError("API_NOT_FOUND", "API 路径不存在。", status=404)


async def dashboard(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    counts = store.dashboard_counts()
    with store.session() as session:
        documents = session.exec(select(KnowledgeDocument)).all()
        counts["rules"] = sum(len(_json_loads(item.structured_json, {}).get("rules", [])) for item in documents)
        running_jobs = session.exec(
            select(func.count()).select_from(Job).where(Job.status.in_(["QUEUED", "RUNNING"]))
        ).one()
    return web.json_response({"metrics": counts | {"running_jobs": int(running_jobs)}, "evaluation_status": "待评测"})


async def list_scenes(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    query_text = request.query.get("q", "").strip().lower()
    status = request.query.get("status", "ALL").upper()
    with store.session() as session:
        statement = select(Scene).where(Scene.archived_at.is_(None)).order_by(Scene.updated_at.desc())
        if status != "ALL":
            statement = statement.where(Scene.status == status)
        rows = session.exec(statement).all()
        items = []
        for scene in rows:
            if query_text and query_text not in f"{scene.name} {scene.description} {scene.goal}".lower():
                continue
            latest = session.exec(
                select(ExtractionRound)
                .where(ExtractionRound.scene_id == scene.id)
                .order_by(ExtractionRound.version.desc())
            ).first()
            material_count = 0
            rule_count = 0
            asset_count = 0
            if latest:
                material_count = int(
                    session.exec(
                        select(func.count()).select_from(Material).where(Material.round_id == latest.id, Material.enabled == True)  # noqa: E712
                    ).one()
                )
                document = session.exec(
                    select(KnowledgeDocument).where(KnowledgeDocument.round_id == latest.id)
                ).first()
                rule_count = len(_json_loads(document.structured_json, {}).get("rules", [])) if document else 0
                asset_count = len(
                    {
                        asset.kind
                        for asset in session.exec(select(Asset).where(Asset.round_id == latest.id)).all()
                    }
                )
            items.append(
                {
                    "id": scene.id,
                    "name": scene.name,
                    "description": scene.description,
                    "goal": scene.goal,
                    "owner": scene.owner,
                    "status": scene.status,
                    "round": _round_dict(latest) if latest else None,
                    "material_count": material_count,
                    "rule_count": rule_count,
                    "asset_count": asset_count,
                    "created_at": _iso(scene.created_at),
                    "updated_at": _iso(scene.updated_at),
                }
            )
    return web.json_response({"items": items})


async def create_scene(request: web.Request) -> web.Response:
    scene, round_row = request.app[APP_SERVICE].create_scene(await _json_body(request))
    return web.json_response({"scene": {"id": scene.id, "name": scene.name}, "round": _round_dict(round_row)}, status=201)


async def get_scene(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    scene_id = request.match_info["scene_id"]
    with store.session() as session:
        scene = session.get(Scene, scene_id)
        if scene is None or scene.archived_at is not None:
            raise WorkbenchError("SCENE_NOT_FOUND", "场景不存在。", status=404)
        rounds = session.exec(
            select(ExtractionRound).where(ExtractionRound.scene_id == scene_id).order_by(ExtractionRound.version.desc())
        ).all()
    return web.json_response(
        {
            "id": scene.id,
            "name": scene.name,
            "description": scene.description,
            "goal": scene.goal,
            "owner": scene.owner,
            "status": scene.status,
            "rounds": [_round_dict(item) for item in rounds],
            "created_at": _iso(scene.created_at),
            "updated_at": _iso(scene.updated_at),
        }
    )


async def update_scene(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    payload = await _json_body(request)
    scene_id = request.match_info["scene_id"]
    with store.session() as session:
        scene = session.get(Scene, scene_id)
        if scene is None or scene.archived_at is not None:
            raise WorkbenchError("SCENE_NOT_FOUND", "场景不存在。", status=404)
        latest = session.exec(
            select(ExtractionRound)
            .where(ExtractionRound.scene_id == scene_id)
            .order_by(ExtractionRound.version.desc())
        ).first()
        if latest and latest.status == "PUBLISHED":
            raise WorkbenchError("ROUND_IMMUTABLE", "已发布轮次不可修改，请创建新一轮。", status=409)
        if "name" in payload:
            name = str(payload["name"]).strip()
            if not name:
                raise WorkbenchError("SCENE_NAME_REQUIRED", "请输入场景名称。", status=422)
            scene.name = name[:160]
        if "description" in payload:
            scene.description = str(payload["description"]).strip()
        if "goal" in payload:
            scene.goal = str(payload["goal"]).strip()
        if latest and "subscenes" in payload:
            latest.subscenes_json = json.dumps(payload["subscenes"], ensure_ascii=False)
            latest.updated_at = utc_now()
            session.add(latest)
        scene.updated_at = utc_now()
        session.add(scene)
        response = {"id": scene.id, "name": scene.name, "updated_at": _iso(scene.updated_at)}
        session.commit()
    return web.json_response(response)


async def archive_scene(request: web.Request) -> web.Response:
    request.app[APP_SERVICE].archive_scene(request.match_info["scene_id"])
    return web.json_response({"archived": True})


async def list_rounds(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    scene_id = request.match_info["scene_id"]
    store.get(Scene, scene_id, code="SCENE_NOT_FOUND")
    with store.session() as session:
        rounds = session.exec(
            select(ExtractionRound).where(ExtractionRound.scene_id == scene_id).order_by(ExtractionRound.version.desc())
        ).all()
    return web.json_response({"items": [_round_dict(item) for item in rounds]})


async def create_round(request: web.Request) -> web.Response:
    round_row = request.app[APP_SERVICE].create_next_round(request.match_info["scene_id"])
    return web.json_response(_round_dict(round_row), status=201)


async def list_materials(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    round_id = request.match_info["round_id"]
    store.get(ExtractionRound, round_id, code="ROUND_NOT_FOUND")
    with store.session() as session:
        rows = session.exec(select(Material).where(Material.round_id == round_id).order_by(Material.created_at)).all()
    return web.json_response({"items": [_material_dict(item) for item in rows]})


async def _receive_upload(
    request: web.Request,
    *,
    round_id: str | None = None,
    exploration_id: str | None = None,
) -> Material:
    settings = request.app[APP_SETTINGS]
    store = request.app[APP_STORE]
    if round_id:
        round_row = store.get(ExtractionRound, round_id, code="ROUND_NOT_FOUND")
        if round_row.status == "PUBLISHED":
            raise WorkbenchError("ROUND_IMMUTABLE", "已发布轮次不可上传素材。", status=409)
        parent = settings.upload_dir / "rounds" / round_id
    else:
        exploration = store.get(Exploration, str(exploration_id), code="EXPLORATION_NOT_FOUND")
        if exploration.status == "ANALYZING":
            raise WorkbenchError("EXPLORATION_JOB_CONFLICT", "探索正在分析，暂不能上传素材。", status=409)
        parent = settings.upload_dir / "explorations" / str(exploration_id)
    parent.mkdir(parents=True, exist_ok=True)
    try:
        reader = await request.multipart()
    except (AssertionError, web.HTTPBadRequest) as exc:
        raise WorkbenchError("UPLOAD_MULTIPART_REQUIRED", "请使用 multipart/form-data 上传文件。", status=400) from exc
    part = await reader.next()
    while part is not None and not part.filename:
        part = await reader.next()
    if part is None or not part.filename:
        raise WorkbenchError("UPLOAD_FILE_REQUIRED", "请选择要上传的文件。", status=422)
    original_name = part.filename
    safe_name = PurePath(original_name).name
    if not safe_name or safe_name != original_name.replace("\\", "/").split("/")[-1] or ".." in PurePath(safe_name).parts:
        raise WorkbenchError("UPLOAD_FILENAME_INVALID", "文件名不安全。", status=422)
    suffix = Path(safe_name).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise WorkbenchError(
            "MATERIAL_FORMAT_UNSUPPORTED",
            "仅支持 PDF、DOCX、XLSX、CSV、TSV、TXT 和 MD。",
            status=415,
            details={"extension": suffix},
        )
    material_id = new_id()
    target_dir = parent / material_id
    target_dir.mkdir(parents=True, exist_ok=False)
    target = target_dir / safe_name
    partial = target_dir / f"{safe_name}.uploading"
    digest = hashlib.sha256()
    size = 0
    try:
        async with aiofiles.open(partial, "wb") as stream:
            while chunk := await part.read_chunk(size=1024 * 1024):
                size += len(chunk)
                if size > settings.max_upload_bytes:
                    raise WorkbenchError(
                        "MATERIAL_TOO_LARGE",
                        f"单文件不能超过 {settings.max_upload_bytes // (1024 * 1024)} MB。",
                        status=413,
                    )
                digest.update(chunk)
                await stream.write(chunk)
        partial.replace(target)
        text = await parse_material(target, material_id)
        parsed_path = target_dir / "parsed.txt"
        async with aiofiles.open(parsed_path, "w", encoding="utf-8") as stream:
            await stream.write(text)
    except Exception:
        shutil.rmtree(target_dir, ignore_errors=True)
        raise
    material = Material(
        id=material_id,
        round_id=round_id,
        exploration_id=exploration_id,
        name=safe_name,
        role=request.query.get("role", "REFERENCE")[:32],
        file_path=str(target),
        parsed_path=str(parsed_path),
        extension=suffix,
        size_bytes=size,
        sha256=digest.hexdigest(),
    )
    with store.session() as session:
        session.add(material)
        session.commit()
        session.refresh(material)
        session.expunge(material)
    return material


async def upload_round_material(request: web.Request) -> web.Response:
    material = await _receive_upload(request, round_id=request.match_info["round_id"])
    return web.json_response(_material_dict(material), status=201)


async def update_material(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    payload = await _json_body(request)
    material_id = request.match_info["material_id"]
    with store.session() as session:
        material = session.get(Material, material_id)
        if material is None:
            raise WorkbenchError("MATERIAL_NOT_FOUND", "素材不存在。", status=404)
        if material.round_id:
            round_row = session.get(ExtractionRound, material.round_id)
            if round_row and round_row.status == "PUBLISHED":
                raise WorkbenchError("ROUND_IMMUTABLE", "已发布轮次不可修改素材。", status=409)
        if "enabled" in payload:
            material.enabled = bool(payload["enabled"])
        if "role" in payload:
            material.role = str(payload["role"])[:32]
        session.add(material)
        session.commit()
        session.refresh(material)
        session.expunge(material)
    return web.json_response(_material_dict(material))


async def create_exploration(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    exploration = Exploration(
        name=str(payload.get("name", "场景探索")).strip()[:160] or "场景探索",
        goal=str(payload.get("goal", "")).strip(),
    )
    store = request.app[APP_STORE]
    with store.session() as session:
        session.add(exploration)
        session.commit()
        session.refresh(exploration)
        session.expunge(exploration)
    return web.json_response(
        {"id": exploration.id, "name": exploration.name, "goal": exploration.goal, "status": exploration.status},
        status=201,
    )


async def list_explorations(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    with store.session() as session:
        rows = session.exec(
            select(Exploration).where(Exploration.archived_at.is_(None)).order_by(Exploration.updated_at.desc())
        ).all()
        items = []
        for row in rows:
            material_count = session.exec(
                select(func.count()).select_from(Material).where(Material.exploration_id == row.id)
            ).one()
            candidate_count = session.exec(
                select(func.count()).select_from(ExplorationCandidate).where(
                    ExplorationCandidate.exploration_id == row.id
                )
            ).one()
            items.append(
                {
                    "id": row.id,
                    "name": row.name,
                    "goal": row.goal,
                    "status": row.status,
                    "material_count": int(material_count),
                    "candidate_count": int(candidate_count),
                    "created_at": _iso(row.created_at),
                    "updated_at": _iso(row.updated_at),
                }
            )
    return web.json_response({"items": items})


async def archive_exploration(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    exploration_id = request.match_info["exploration_id"]
    with store.session() as session:
        exploration = session.get(Exploration, exploration_id)
        if exploration is None or exploration.archived_at is not None:
            raise WorkbenchError("EXPLORATION_NOT_FOUND", "探索记录不存在。", status=404)
        running = session.exec(
            select(Job).where(Job.exploration_id == exploration_id, Job.status.in_(["QUEUED", "RUNNING"]))
        ).first()
        if running:
            raise WorkbenchError(
                "EXPLORATION_JOB_CONFLICT",
                "探索任务运行中，完成或失败后才能归档。",
                status=409,
                retryable=True,
                details={"job_id": running.id},
            )
        exploration.status = "ARCHIVED"
        exploration.archived_at = utc_now()
        exploration.updated_at = utc_now()
        session.add(exploration)
        session.commit()
    return web.json_response({"archived": True})


async def upload_exploration_material(request: web.Request) -> web.Response:
    material = await _receive_upload(request, exploration_id=request.match_info["exploration_id"])
    return web.json_response(_material_dict(material), status=201)


async def analyze_exploration(request: web.Request) -> web.Response:
    job = request.app[APP_SERVICE].start_exploration(request.match_info["exploration_id"])
    return web.json_response(_job_dict(job), status=202)


async def list_exploration_materials(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    exploration_id = request.match_info["exploration_id"]
    store.get(Exploration, exploration_id, code="EXPLORATION_NOT_FOUND")
    with store.session() as session:
        rows = session.exec(
            select(Material).where(Material.exploration_id == exploration_id).order_by(Material.created_at)
        ).all()
    return web.json_response({"items": [_material_dict(item) for item in rows]})


async def list_candidates(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    exploration_id = request.match_info["exploration_id"]
    store.get(Exploration, exploration_id, code="EXPLORATION_NOT_FOUND")
    with store.session() as session:
        rows = session.exec(
            select(ExplorationCandidate)
            .where(ExplorationCandidate.exploration_id == exploration_id)
            .order_by(ExplorationCandidate.confidence.desc())
        ).all()
    return web.json_response(
        {
            "items": [
                {
                    "id": row.id,
                    "name": row.name,
                    "description": row.description,
                    "goal": row.goal,
                    "confidence": row.confidence,
                    "source_refs": _json_loads(row.source_refs_json, []),
                    "created_scene_id": row.created_scene_id,
                }
                for row in rows
            ]
        }
    )


async def candidate_create_scene(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    service = request.app[APP_SERVICE]
    exploration_id = request.match_info["exploration_id"]
    candidate_id = request.match_info["candidate_id"]
    with store.session() as session:
        candidate = session.get(ExplorationCandidate, candidate_id)
        if candidate is None or candidate.exploration_id != exploration_id:
            raise WorkbenchError("CANDIDATE_NOT_FOUND", "候选场景不存在。", status=404)
        if candidate.created_scene_id:
            existing_id = candidate.created_scene_id
            return web.json_response({"scene_id": existing_id, "already_created": True})
        materials = session.exec(
            select(Material).where(Material.exploration_id == exploration_id, Material.enabled == True)  # noqa: E712
        ).all()
        for material in materials:
            session.expunge(material)
    scene, round_row = service.create_scene(
        {"name": candidate.name, "description": candidate.description, "goal": candidate.goal}
    )
    with store.session() as session:
        for material in materials:
            session.add(
                Material(
                    round_id=round_row.id,
                    name=material.name,
                    role=material.role,
                    file_path=material.file_path,
                    parsed_path=material.parsed_path,
                    extension=material.extension,
                    size_bytes=material.size_bytes,
                    sha256=material.sha256,
                )
            )
        candidate_row = session.get(ExplorationCandidate, candidate_id)
        if candidate_row:
            candidate_row.created_scene_id = scene.id
            session.add(candidate_row)
        session.commit()
    return web.json_response({"scene_id": scene.id, "round_id": round_row.id}, status=201)


async def start_extraction(request: web.Request) -> web.Response:
    job = request.app[APP_SERVICE].start_extraction(request.match_info["round_id"])
    return web.json_response(_job_dict(job), status=202)


async def get_job(request: web.Request) -> web.Response:
    job = request.app[APP_STORE].get(Job, request.match_info["job_id"], code="JOB_NOT_FOUND")
    return web.json_response(_job_dict(job))


async def job_events(request: web.Request) -> web.StreamResponse:
    store = request.app[APP_STORE]
    job_id = request.match_info["job_id"]
    store.get(Job, job_id, code="JOB_NOT_FOUND")
    raw_last = request.headers.get("Last-Event-ID", request.query.get("after", "0"))
    try:
        last_seq = max(0, int(raw_last))
    except ValueError:
        last_seq = 0
    response = web.StreamResponse(
        status=200,
        headers={
            "Content-Type": "text/event-stream",
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )
    await response.prepare(request)
    try:
        while True:
            rows = store.events_after(job_id, last_seq)
            for row in rows:
                payload = {
                    "seq": row.seq,
                    "phase": row.phase,
                    "status": row.status,
                    "progress": row.progress,
                    "message": row.message,
                }
                await response.write(
                    f"id: {row.seq}\nevent: progress\ndata: {json.dumps(payload, ensure_ascii=False)}\n\n".encode()
                )
                last_seq = row.seq
            job = store.get(Job, job_id, code="JOB_NOT_FOUND")
            if job.status in TERMINAL_JOB_STATUSES and not rows:
                break
            await asyncio.sleep(0.15)
    except (ConnectionResetError, asyncio.CancelledError):
        pass
    return response


async def retry_job(request: web.Request) -> web.Response:
    job = request.app[APP_SERVICE].retry_job(request.match_info["job_id"])
    return web.json_response(_job_dict(job), status=202)


async def get_document(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    round_id = request.match_info["round_id"]
    store.get(ExtractionRound, round_id, code="ROUND_NOT_FOUND")
    with store.session() as session:
        document = session.exec(select(KnowledgeDocument).where(KnowledgeDocument.round_id == round_id)).first()
    if document is None:
        raise WorkbenchError("DOCUMENT_NOT_FOUND", "当前轮次尚未生成知识文档。", status=404)
    return web.json_response(_document_dict(document))


async def save_document(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    round_id = request.match_info["round_id"]
    payload = await _json_body(request)
    markdown = str(payload.get("markdown", ""))
    try:
        base_revision = int(payload.get("base_revision"))
    except (TypeError, ValueError) as exc:
        raise WorkbenchError("BASE_REVISION_REQUIRED", "保存时必须提供 base_revision。", status=422) from exc
    with store.session() as session:
        round_row = session.get(ExtractionRound, round_id)
        if round_row is None:
            raise WorkbenchError("ROUND_NOT_FOUND", "萃取轮次不存在。", status=404)
        if round_row.status == "PUBLISHED":
            raise WorkbenchError("ROUND_IMMUTABLE", "已发布轮次不可修改。", status=409)
        document = session.exec(select(KnowledgeDocument).where(KnowledgeDocument.round_id == round_id)).first()
        if document is None:
            raise WorkbenchError("DOCUMENT_NOT_FOUND", "知识文档不存在。", status=404)
        if document.revision != base_revision:
            raise WorkbenchError(
                "DOCUMENT_REVISION_CONFLICT",
                "文档已被更新，请刷新后再保存。",
                status=409,
                details={"current_revision": document.revision},
            )
        document.markdown = markdown
        document.revision += 1
        document.updated_at = utc_now()
        session.add(document)
        session.add(
            Revision(
                document_id=document.id,
                revision=document.revision,
                markdown=markdown,
                reason=str(payload.get("reason", "手工保存"))[:200],
                author="本机用户",
            )
        )
        session.commit()
        session.refresh(document)
        session.expunge(document)
    return web.json_response(_document_dict(document))


async def list_revisions(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    round_id = request.match_info["round_id"]
    with store.session() as session:
        document = session.exec(select(KnowledgeDocument).where(KnowledgeDocument.round_id == round_id)).first()
        if document is None:
            return web.json_response({"items": []})
        rows = session.exec(
            select(Revision).where(Revision.document_id == document.id).order_by(Revision.revision.desc())
        ).all()
    return web.json_response(
        {
            "items": [
                {
                    "id": row.id,
                    "revision": row.revision,
                    "reason": row.reason,
                    "author": row.author,
                    "created_at": _iso(row.created_at),
                }
                for row in rows
            ]
        }
    )


async def list_suggestions(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    round_id = request.match_info["round_id"]
    with store.session() as session:
        rows = session.exec(
            select(Suggestion).where(Suggestion.round_id == round_id).order_by(Suggestion.created_at.desc())
        ).all()
    return web.json_response({"items": [_suggestion_dict(item) for item in rows]})


async def create_suggestion(request: web.Request) -> web.Response:
    payload = await _json_body(request) if request.can_read_body else {}
    suggestion = await request.app[APP_SERVICE].create_suggestion(
        request.match_info["round_id"],
        mode=str(payload.get("mode", "CONSISTENCY")),
        instruction=str(payload.get("instruction", "")),
    )
    return web.json_response(_suggestion_dict(suggestion), status=201)


async def apply_suggestion(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    try:
        base_revision = int(payload.get("base_revision"))
    except (TypeError, ValueError) as exc:
        raise WorkbenchError("BASE_REVISION_REQUIRED", "采纳建议时必须提供 base_revision。", status=422) from exc
    document = request.app[APP_SERVICE].apply_suggestion(
        request.match_info["suggestion_id"], expected_revision=base_revision
    )
    return web.json_response(_document_dict(document))


async def reject_suggestion(request: web.Request) -> web.Response:
    suggestion = request.app[APP_SERVICE].reject_suggestion(request.match_info["suggestion_id"])
    return web.json_response(_suggestion_dict(suggestion))


async def list_assets(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    round_id = request.match_info["round_id"]
    store.get(ExtractionRound, round_id, code="ROUND_NOT_FOUND")
    with store.session() as session:
        rows = session.exec(
            select(Asset).where(Asset.round_id == round_id).order_by(Asset.kind, Asset.version.desc())
        ).all()
        document = session.exec(select(KnowledgeDocument).where(KnowledgeDocument.round_id == round_id)).first()
    latest: dict[str, Asset] = {}
    for row in rows:
        latest.setdefault(row.kind, row)
    return web.json_response(
        {
            "items": [_asset_dict(item, document.revision if document else None) for item in latest.values()],
            "complete": bool(document)
            and REQUIRED_ASSET_KINDS.issubset(latest.keys())
            and all(item.source_revision == document.revision for item in latest.values()),
        }
    )


async def generate_round_assets(request: web.Request) -> web.Response:
    payload = await _json_body(request) if request.can_read_body else {}
    job = request.app[APP_SERVICE].start_asset_generation(
        request.match_info["round_id"], only_kind=payload.get("kind")
    )
    return web.json_response(_job_dict(job), status=202)


async def download_asset(request: web.Request) -> web.StreamResponse:
    asset = request.app[APP_STORE].get(Asset, request.match_info["asset_id"], code="ASSET_NOT_FOUND")
    path = Path(asset.file_path)
    if not path.is_file():
        raise WorkbenchError("ASSET_FILE_MISSING", "资产文件不存在，请重新生成。", status=404)
    response = web.FileResponse(path)
    response.content_type = asset.mime_type
    response.headers["Content-Disposition"] = f'attachment; filename="{asset.filename}"'
    return response


async def preview_asset(request: web.Request) -> web.Response:
    asset = request.app[APP_STORE].get(Asset, request.match_info["asset_id"], code="ASSET_NOT_FOUND")
    path = Path(asset.file_path)
    if not path.is_file():
        raise WorkbenchError("ASSET_FILE_MISSING", "资产文件不存在，请重新生成。", status=404)

    if asset.kind == "RULES_XLSX":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values: list[list[str | int | float | bool | None]] = []
            truncated = False
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index > 50:
                    truncated = True
                    break
                values.append([_preview_cell(value) for value in row])
        finally:
            workbook.close()
        payload = _asset_base_preview(asset, "table") | {
            "sheet": sheet.title,
            "columns": [str(value or "") for value in values[0]] if values else [],
            "rows": values[1:] if values else [],
            "truncated": truncated,
        }
    elif asset.kind == "THOUGHT_CHAIN_MD":
        with path.open("r", encoding="utf-8", errors="replace") as file:
            text = file.read(20_001)
        payload = _asset_base_preview(asset, "markdown") | {
            "text": text[:20_000],
            "truncated": len(text) > 20_000,
        }
    elif asset.kind == "SKILL_ZIP":
        with zipfile.ZipFile(path) as archive:
            infos = [info for info in archive.infolist() if not info.is_dir()]
            entries = [
                {"path": info.filename, "size_bytes": info.file_size}
                for info in infos[:100]
            ]
            skill_text = ""
            try:
                skill_info = archive.getinfo("SKILL.md")
                if skill_info.file_size <= 64 * 1024:
                    skill_text = archive.read(skill_info).decode("utf-8", errors="replace")
            except KeyError:
                pass
        payload = _asset_base_preview(asset, "archive") | {
            "entries": entries,
            "text": skill_text,
            "truncated": len(infos) > 100,
        }
    elif asset.kind in {"QA_JSONL", "EVAL_JSONL"}:
        items: list[Any] = []
        truncated = False
        with path.open("r", encoding="utf-8", errors="replace") as file:
            for line in file:
                stripped = line.strip()
                if not stripped:
                    continue
                if len(items) >= 50:
                    truncated = True
                    break
                try:
                    items.append(json.loads(stripped))
                except json.JSONDecodeError:
                    items.append({"raw": stripped})
        payload = _asset_base_preview(asset, "jsonl") | {
            "items": items,
            "truncated": truncated,
        }
    else:
        raise WorkbenchError("ASSET_PREVIEW_UNSUPPORTED", "该资产类型暂不支持预览。", status=422)

    response = web.json_response(payload)
    response.headers["Cache-Control"] = "no-store"
    return response


async def download_round_assets(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    round_id = request.match_info["round_id"]
    round_row = store.get(ExtractionRound, round_id, code="ROUND_NOT_FOUND")
    with store.session() as session:
        rows = session.exec(
            select(Asset).where(Asset.round_id == round_id).order_by(Asset.kind, Asset.version.desc())
        ).all()
        document = session.exec(select(KnowledgeDocument).where(KnowledgeDocument.round_id == round_id)).first()
    if document is None:
        raise WorkbenchError("DOCUMENT_REQUIRED", "知识文档不存在。", status=409)

    latest: dict[str, Asset] = {}
    for row in rows:
        latest.setdefault(row.kind, row)
    unavailable = sorted(
        kind
        for kind in REQUIRED_ASSET_KINDS
        if kind not in latest or latest[kind].source_revision != document.revision
    )
    if unavailable:
        raise WorkbenchError(
            "ASSETS_INCOMPLETE",
            "五类当前版本资产尚未全部生成，无法打包下载。",
            status=409,
            details={"unavailable": unavailable, "document_revision": document.revision},
        )

    assets = [latest[kind] for kind in sorted(REQUIRED_ASSET_KINDS)]
    missing_files = [asset.kind for asset in assets if not Path(asset.file_path).is_file()]
    if missing_files:
        raise WorkbenchError(
            "ASSET_FILE_MISSING",
            "部分资产文件不存在，请重新生成后下载。",
            status=404,
            details={"missing": missing_files},
        )

    manifest = {
        "round_id": round_id,
        "round_version": round_row.version,
        "document_revision": document.revision,
        "assets": [
            {
                "kind": asset.kind,
                "filename": asset.filename,
                "version": asset.version,
                "source_revision": asset.source_revision,
                "synthetic": asset.synthetic,
            }
            for asset in assets
        ],
    }
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
        for asset in assets:
            archive.write(asset.file_path, arcname=asset.filename)

    response = web.Response(body=buffer.getvalue(), content_type="application/zip")
    response.headers["Content-Disposition"] = f'attachment; filename="knowledge-assets-v{round_row.version}.zip"'
    response.headers["Cache-Control"] = "no-store"
    return response


async def publish_round(request: web.Request) -> web.Response:
    round_row = request.app[APP_SERVICE].publish_round(request.match_info["round_id"])
    return web.json_response(_round_dict(round_row))


def _published_skill_dict(
    round_row: ExtractionRound,
    scene: Scene,
    skill_asset: Asset | None,
    evaluation_asset: Asset | None,
) -> dict[str, Any]:
    return {
        "id": round_row.id,
        "round_id": round_row.id,
        "scene_id": scene.id,
        "name": scene.name,
        "version": round_row.version,
        "label": f"{scene.name} v{round_row.version}",
        "published_at": _iso(round_row.published_at),
        "has_skill_asset": skill_asset is not None,
        "evaluation_asset": _asset_dict(evaluation_asset) if evaluation_asset else None,
    }


async def list_runtime_skills(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    with store.session() as session:
        rounds = session.exec(
            select(ExtractionRound)
            .where(ExtractionRound.status == "PUBLISHED")
            .order_by(ExtractionRound.published_at.desc())
        ).all()
        items = []
        for round_row in rounds:
            scene = session.get(Scene, round_row.scene_id)
            if scene is None or scene.archived_at is not None:
                continue
            skill_asset = session.exec(
                select(Asset)
                .where(Asset.round_id == round_row.id, Asset.kind == "SKILL_ZIP")
                .order_by(Asset.version.desc())
            ).first()
            evaluation_asset = session.exec(
                select(Asset)
                .where(Asset.round_id == round_row.id, Asset.kind == "EVAL_JSONL")
                .order_by(Asset.version.desc())
            ).first()
            items.append(_published_skill_dict(round_row, scene, skill_asset, evaluation_asset))
    return web.json_response({"items": items})


async def _receive_form_file(
    request: web.Request,
    *,
    parent: Path,
    allowed_extensions: set[str],
) -> tuple[dict[str, str], Path, str, str]:
    settings = request.app[APP_SETTINGS]
    try:
        reader = await request.multipart()
    except (AssertionError, web.HTTPBadRequest) as exc:
        raise WorkbenchError("UPLOAD_MULTIPART_REQUIRED", "请使用 multipart/form-data 上传文件。", status=400) from exc
    fields: dict[str, str] = {}
    target: Path | None = None
    safe_name = ""
    digest_hex = ""
    part = await reader.next()
    try:
        while part is not None:
            if part.filename and target is None:
                original_name = part.filename
                safe_name = PurePath(original_name).name
                if not safe_name or safe_name != original_name.replace("\\", "/").split("/")[-1]:
                    raise WorkbenchError("UPLOAD_FILENAME_INVALID", "文件名不安全。", status=422)
                suffix = Path(safe_name).suffix.lower()
                if suffix not in allowed_extensions:
                    raise WorkbenchError(
                        "DATASET_FORMAT_UNSUPPORTED",
                        "文件格式不受支持，请按页面提示上传。",
                        status=415,
                        details={"extension": suffix},
                    )
                parent.mkdir(parents=True, exist_ok=False)
                target = parent / safe_name
                partial = parent / f"{safe_name}.uploading"
                digest = hashlib.sha256()
                size = 0
                async with aiofiles.open(partial, "wb") as stream:
                    while chunk := await part.read_chunk(size=1024 * 1024):
                        size += len(chunk)
                        if size > settings.max_upload_bytes:
                            raise WorkbenchError("DATASET_TOO_LARGE", "上传文件超过单文件大小限制。", status=413)
                        digest.update(chunk)
                        await stream.write(chunk)
                partial.replace(target)
                digest_hex = digest.hexdigest()
            elif part.name:
                fields[part.name] = (await part.text()).strip()
            part = await reader.next()
    except Exception:
        shutil.rmtree(parent, ignore_errors=True)
        raise
    if target is None:
        raise WorkbenchError("UPLOAD_FILE_REQUIRED", "请选择要上传的文件。", status=422)
    return fields, target, safe_name, digest_hex


def _rows_from_file(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".jsonl":
        rows = []
        for line in path.read_text(encoding="utf-8-sig").splitlines():
            if not line.strip():
                continue
            value = json.loads(line)
            if isinstance(value, dict):
                rows.append(value)
        return rows
    if suffix == ".json":
        value = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(value, dict):
            value = value.get("items", value.get("cases", []))
        return [item for item in value if isinstance(item, dict)] if isinstance(value, list) else []
    if suffix in {".csv", ".tsv"}:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            return list(csv.DictReader(stream, delimiter="\t" if suffix == ".tsv" else ","))
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values, [])]
            return [
                {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
                for row in values
                if any(value not in (None, "") for value in row)
            ]
        finally:
            workbook.close()
    if suffix in {".txt", ".md"}:
        return [
            {"input": line.strip(), "original_output": ""}
            for line in path.read_text(encoding="utf-8-sig").splitlines()
            if line.strip()
        ]
    return []


def _first_value(row: dict[str, Any], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def _evaluation_cases(path: Path) -> list[dict[str, Any]]:
    try:
        rows = _rows_from_file(path)
    except (json.JSONDecodeError, UnicodeDecodeError, ValueError) as exc:
        raise WorkbenchError("EVALUATION_DATASET_INVALID", "测试集无法解析。", status=422) from exc
    cases = []
    for index, row in enumerate(rows[:101], start=1):
        input_text = _first_value(row, ("input", "输入", "question", "问题"))
        expected = _first_value(row, ("expected", "标准答案", "answer", "答案", "label", "标签"))
        if input_text and expected:
            cases.append(
                {
                    "id": _first_value(row, ("id", "编号")) or f"C-{index:03d}",
                    "input": input_text[:20000],
                    "expected": expected[:4000],
                    "source_refs": row.get("source_refs", []) if isinstance(row.get("source_refs"), list) else [],
                }
            )
    if not cases:
        raise WorkbenchError(
            "EVALUATION_DATASET_EMPTY",
            "测试集至少需要 input/expected（或输入/标准答案）两列。",
            status=422,
        )
    return cases


def _feedback_cases(path: Path) -> list[dict[str, Any]]:
    try:
        rows = _rows_from_file(path)
    except (json.JSONDecodeError, UnicodeDecodeError, ValueError) as exc:
        raise WorkbenchError("FEEDBACK_DATASET_INVALID", "错例文件无法解析。", status=422) from exc
    cases = []
    for index, row in enumerate(rows[:101], start=1):
        input_text = _first_value(row, ("input", "输入", "question", "问题", "case", "案例"))
        original_output = _first_value(row, ("original_output", "原输出", "output", "错误结果", "模型输出"))
        expected = _first_value(row, ("expected", "标准答案", "correct", "正确结果", "正确标签"))
        if input_text:
            cases.append(
                {
                    "id": _first_value(row, ("id", "编号")) or f"E-{index:03d}",
                    "summary": _first_value(row, ("summary", "摘要", "name", "名称")) or input_text[:80],
                    "input": input_text[:20000],
                    "original_output": original_output[:10000],
                    "expected": expected[:4000],
                }
            )
    if not cases:
        raise WorkbenchError("FEEDBACK_CASES_EMPTY", "错例文件中没有可用输入。", status=422)
    return cases


async def runtime_tryout(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    result = await request.app[APP_SERVICE].run_tryout(
        str(payload.get("round_id", "")),
        str(payload.get("model_connection_id", "")),
        str(payload.get("input", "")),
    )
    return web.json_response(result)


async def runtime_tryout_upload(request: web.Request) -> web.Response:
    parent = request.app[APP_SETTINGS].upload_dir / "runtime" / new_id()
    try:
        fields, path, _, _ = await _receive_form_file(
            request,
            parent=parent,
            allowed_extensions=set(SUPPORTED_EXTENSIONS),
        )
        parsed = await parse_material(path, new_id())
        instruction = fields.get("input", "")
        input_text = f"用户说明：{instruction}\n\n上传资料：\n{parsed}" if instruction else parsed
        result = await request.app[APP_SERVICE].run_tryout(
            fields.get("round_id", ""),
            fields.get("model_connection_id", ""),
            input_text,
        )
        return web.json_response(result)
    finally:
        shutil.rmtree(parent, ignore_errors=True)


async def list_evaluations(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    statement = select(EvaluationRun).order_by(EvaluationRun.created_at.desc())
    round_id = request.query.get("round_id", "").strip()
    if round_id:
        statement = statement.where(EvaluationRun.round_id == round_id)
    with store.session() as session:
        rows = session.exec(statement).all()
    return web.json_response({"items": [_evaluation_dict(item) for item in rows]})


async def get_evaluation(request: web.Request) -> web.Response:
    row = request.app[APP_STORE].get(EvaluationRun, request.match_info["evaluation_id"], code="EVALUATION_NOT_FOUND")
    return web.json_response(_evaluation_dict(row))


async def create_evaluation(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    round_id = str(payload.get("round_id", ""))
    store = request.app[APP_STORE]
    with store.session() as session:
        asset = session.exec(
            select(Asset)
            .where(Asset.round_id == round_id, Asset.kind == "EVAL_JSONL")
            .order_by(Asset.version.desc())
        ).first()
        if asset:
            session.expunge(asset)
    if asset is None or not Path(asset.file_path).is_file():
        raise WorkbenchError("EVALUATION_ASSET_REQUIRED", "该 Skill 尚无可用评测集资产。", status=409)
    cases = _evaluation_cases(Path(asset.file_path))
    evaluation, job = request.app[APP_SERVICE].start_evaluation(
        round_id=round_id,
        model_id=str(payload.get("model_connection_id", "")),
        dataset_name=asset.filename,
        dataset_kind="GENERATED",
        dataset_path=asset.file_path,
        dataset_sha256=hashlib.sha256(Path(asset.file_path).read_bytes()).hexdigest(),
        cases=cases,
    )
    return web.json_response({"evaluation": _evaluation_dict(evaluation), "job": _job_dict(job)}, status=202)


async def upload_evaluation(request: web.Request) -> web.Response:
    parent = request.app[APP_SETTINGS].upload_dir / "evaluations" / new_id()
    fields, path, filename, digest = await _receive_form_file(
        request,
        parent=parent,
        allowed_extensions={".jsonl", ".csv", ".tsv", ".xlsx"},
    )
    cases = _evaluation_cases(path)
    evaluation, job = request.app[APP_SERVICE].start_evaluation(
        round_id=fields.get("round_id", ""),
        model_id=fields.get("model_connection_id", ""),
        dataset_name=filename,
        dataset_kind="UPLOADED",
        dataset_path=str(path),
        dataset_sha256=digest,
        cases=cases,
    )
    return web.json_response({"evaluation": _evaluation_dict(evaluation), "job": _job_dict(job)}, status=202)


async def evaluation_to_feedback(request: web.Request) -> web.Response:
    evaluation = request.app[APP_STORE].get(
        EvaluationRun, request.match_info["evaluation_id"], code="EVALUATION_NOT_FOUND"
    )
    if evaluation.status != "COMPLETED":
        raise WorkbenchError("EVALUATION_NOT_COMPLETE", "评测完成后才能回流错例。", status=409)
    payload = await _json_body(request)
    results = _json_loads(evaluation.results_json, [])
    cases = [
        {
            "id": item.get("id"),
            "summary": str(item.get("input", ""))[:80],
            "input": item.get("input", ""),
            "original_output": item.get("answer", ""),
            "expected": item.get("expected", ""),
        }
        for item in results
        if isinstance(item, dict) and item.get("correct") is False
    ]
    if not cases:
        raise WorkbenchError("EVALUATION_ERRORS_EMPTY", "本次评测没有可回流错例。", status=409)
    task = request.app[APP_SERVICE].create_feedback_task(
        round_id=evaluation.round_id,
        model_id=evaluation.model_connection_id,
        name=str(payload.get("name") or f"{evaluation.dataset_name} · 错例分析"),
        task_type=str(payload.get("task_type", "CLASSIFICATION")),
        cases=cases,
        source_filename=evaluation.dataset_name,
    )
    return web.json_response(_feedback_task_dict(task), status=201)


async def list_feedback_tasks(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    with store.session() as session:
        rows = session.exec(select(FeedbackTask).order_by(FeedbackTask.created_at.desc())).all()
    return web.json_response({"items": [_feedback_task_dict(item) for item in rows]})


async def get_feedback_task(request: web.Request) -> web.Response:
    task = request.app[APP_STORE].get(
        FeedbackTask, request.match_info["task_id"], code="FEEDBACK_TASK_NOT_FOUND"
    )
    return web.json_response(_feedback_task_dict(task))


async def create_feedback_task(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    task = request.app[APP_SERVICE].create_feedback_task(
        round_id=str(payload.get("round_id", "")),
        model_id=str(payload.get("model_connection_id", "")),
        name=str(payload.get("name", "")),
        task_type=str(payload.get("task_type", "CLASSIFICATION")),
    )
    return web.json_response(_feedback_task_dict(task), status=201)


async def upload_feedback_cases(request: web.Request) -> web.Response:
    task_id = request.match_info["task_id"]
    parent = request.app[APP_SETTINGS].upload_dir / "feedback" / task_id / new_id()
    _, path, filename, digest = await _receive_form_file(
        request,
        parent=parent,
        allowed_extensions={".json", ".jsonl", ".csv", ".tsv", ".xlsx", ".txt", ".md"},
    )
    task = request.app[APP_SERVICE].replace_feedback_cases(
        task_id,
        _feedback_cases(path),
        source_filename=filename,
        source_path=str(path),
        source_sha256=digest,
    )
    return web.json_response(_feedback_task_dict(task))


async def analyze_feedback_task(request: web.Request) -> web.Response:
    task, job = request.app[APP_SERVICE].start_feedback_analysis(request.match_info["task_id"])
    return web.json_response({"task": _feedback_task_dict(task), "job": _job_dict(job)}, status=202)


async def save_feedback_task(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    cases = payload.get("cases", [])
    if not isinstance(cases, list):
        raise WorkbenchError("FEEDBACK_CASES_INVALID", "cases 必须是数组。", status=422)
    task = request.app[APP_SERVICE].save_feedback_cases(request.match_info["task_id"], cases)
    return web.json_response(_feedback_task_dict(task))


async def export_feedback_task(request: web.Request) -> web.Response:
    task = request.app[APP_STORE].get(
        FeedbackTask, request.match_info["task_id"], code="FEEDBACK_TASK_NOT_FOUND"
    )
    cases = _json_loads(task.cases_json, [])
    output_format = request.query.get("format", "json").lower()
    if output_format == "json":
        body = json.dumps({"task": task.name, "task_type": task.task_type, "items": cases}, ensure_ascii=False, indent=2)
        response = web.Response(text=body, content_type="application/json")
        response.headers["Content-Disposition"] = f'attachment; filename="feedback-{task.id[:8]}.json"'
        return response
    if output_format != "xlsx":
        raise WorkbenchError("FEEDBACK_EXPORT_FORMAT_INVALID", "仅支持 JSON 或 XLSX。", status=422)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "错例分析"
    sheet.append(["编号", "输入", "原输出", "标准答案", "专家结论", "错因/问题", "知识缺口", "归因"])
    for case in cases:
        expert = case.get("expert", {}) if isinstance(case.get("expert"), dict) else {}
        sheet.append(
            [
                case.get("id", ""),
                case.get("input", ""),
                case.get("original_output", ""),
                case.get("expected", ""),
                expert.get("correct_label") or expert.get("expected_content") or "",
                expert.get("error_reason")
                or "\n".join(
                    str(item.get("description", ""))
                    for item in expert.get("issues", [])
                    if isinstance(item, dict)
                ),
                expert.get("knowledge_gap", ""),
                expert.get("attribution", ""),
            ]
        )
    buffer = io.BytesIO()
    workbook.save(buffer)
    response = web.Response(body=buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f'attachment; filename="feedback-{task.id[:8]}.xlsx"'
    return response


async def promote_feedback_task(request: web.Request) -> web.Response:
    task, round_row, scene = request.app[APP_SERVICE].promote_feedback_task(request.match_info["task_id"])
    return web.json_response(
        {"task": _feedback_task_dict(task), "round": _round_dict(round_row), "scene_id": scene.id}
    )


def _validate_model_endpoint(provider: str, api_base: str) -> None:
    if provider == "FakeModel":
        raise WorkbenchError("MODEL_PROVIDER_UNSUPPORTED", "产品运行不再支持 Fake Model。", status=422)
    parsed = urlparse(api_base)
    if parsed.scheme == "https" and parsed.hostname:
        return
    if parsed.scheme == "http" and parsed.hostname:
        configured_hosts = {
            item.strip().lower().strip("[]")
            for item in os.environ.get("WORKBENCH_MODEL_HTTP_HOSTS", "").split(",")
            if item.strip() and item.strip() != "*"
        }
        if parsed.hostname.lower() in {"127.0.0.1", "localhost", "::1", *configured_hosts}:
            return
    raise WorkbenchError(
        "MODEL_API_BASE_INVALID",
        "API 地址必须使用 HTTPS；HTTP 仅允许本机或 WORKBENCH_MODEL_HTTP_HOSTS 中的精确主机。",
        status=422,
    )


async def list_models(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    with store.session() as session:
        rows = session.exec(
            select(ModelConnection)
            .where(ModelConnection.provider != "FakeModel")
            .order_by(ModelConnection.created_at)
        ).all()
    return web.json_response({"items": [_model_dict(item) for item in rows]})


async def create_model(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    name = str(payload.get("name", "")).strip()
    provider = str(payload.get("provider", "")).strip()
    api_base = str(payload.get("api_base", "")).strip()
    model_name = str(payload.get("model_name", "")).strip()
    api_key = str(payload.get("api_key", "")).strip()
    if not name or not provider or not model_name:
        raise WorkbenchError("MODEL_FIELDS_REQUIRED", "名称、调用适配器和模型名称为必填项。", status=422)
    if not api_key:
        raise WorkbenchError("MODEL_API_KEY_REQUIRED", "请输入 API Key。", status=422)
    _validate_model_endpoint(provider, api_base)
    model = ModelConnection(
        name=name[:120],
        provider=provider[:40],
        api_base=api_base[:500],
        model_name=model_name[:160],
        encrypted_api_key=request.app[APP_SECRET_BOX].encrypt(api_key),
        enabled=bool(payload.get("enabled", True)),
    )
    store = request.app[APP_STORE]
    with store.session() as session:
        session.add(model)
        session.commit()
        session.refresh(model)
        session.expunge(model)
    return web.json_response(_model_dict(model), status=201)


async def update_model(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    store = request.app[APP_STORE]
    model_id = request.match_info["model_id"]
    with store.session() as session:
        model = session.get(ModelConnection, model_id)
        if model is None:
            raise WorkbenchError("MODEL_NOT_FOUND", "模型连接不存在。", status=404)
        provider = str(payload.get("provider", model.provider)).strip()
        api_base = str(payload.get("api_base", model.api_base)).strip()
        _validate_model_endpoint(provider, api_base)
        for field, limit in (("name", 120), ("provider", 40), ("api_base", 500), ("model_name", 160)):
            if field in payload:
                setattr(model, field, str(payload[field]).strip()[:limit])
        if payload.get("api_key"):
            model.encrypted_api_key = request.app[APP_SECRET_BOX].encrypt(str(payload["api_key"]).strip())
        if "enabled" in payload:
            model.enabled = bool(payload["enabled"])
        model.updated_at = utc_now()
        session.add(model)
        session.commit()
        session.refresh(model)
        session.expunge(model)
    return web.json_response(_model_dict(model))


async def delete_model(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    model_id = request.match_info["model_id"]
    with store.session() as session:
        model = session.get(ModelConnection, model_id)
        if model is None:
            raise WorkbenchError("MODEL_NOT_FOUND", "模型连接不存在。", status=404)
        mount = session.exec(select(AbilityMount).where(AbilityMount.model_connection_id == model_id)).first()
        profile = session.exec(select(AbilityProfile).where(AbilityProfile.model_connection_id == model_id)).first()
        if mount or profile:
            raise WorkbenchError(
                "MODEL_IN_USE",
                "该模型仍被能力配置使用，请先更换挂载模型。",
                status=409,
                details={"ability_key": mount.ability_key if mount else "SCENE_PROFILE"},
            )
        session.delete(model)
        session.commit()
    return web.json_response({"deleted": True})


async def test_model_connection(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    model = store.get(ModelConnection, request.match_info["model_id"], code="MODEL_NOT_FOUND")
    if not model.encrypted_api_key:
        raise WorkbenchError("MODEL_API_KEY_REQUIRED", "该连接尚未保存 API Key。", status=422)
    _validate_model_endpoint(model.provider, model.api_base)
    started = asyncio.get_running_loop().time()
    try:
        from openjiuwen.core.foundation.llm import Model
        from openjiuwen.core.foundation.llm.schema.config import ModelRequestConfig

        client = build_model_client_config(
            provider=model.provider,
            api_key=request.app[APP_SECRET_BOX].decrypt(model.encrypted_api_key),
            api_base=model.api_base,
            timeout=20,
            max_retries=0,
        )
        runtime = Model(client, ModelRequestConfig(model=model.model_name, temperature=0, max_tokens=256))
        response = await runtime.invoke(
            (
                "只返回合法 JSON，不要输出 Markdown 或解释。严格返回："
                '{"items":[{"question":"何时需要人工复核？","answer":"命中异常条件时。","source_refs":[]}]}。'
            ),
            model=model.model_name,
            timeout=20,
        )
    except Exception as exc:
        raise model_connection_error(exc) from exc
    parsed = OpenJiuwenKnowledgeModel._decode_json(getattr(response, "content", None))
    if not OpenJiuwenKnowledgeModel._normalize_qa_items(parsed):
        raise WorkbenchError(
            "MODEL_STRUCTURED_OUTPUT_INVALID",
            "模型连接成功，但未返回可用的 QA JSON 结构。请检查模型的结构化输出能力。",
            status=422,
            retryable=True,
            details={"response_shape": OpenJiuwenKnowledgeModel._json_shape_summary(parsed)},
        )
    latency_ms = round((asyncio.get_running_loop().time() - started) * 1000)
    return web.json_response({"ok": True, "latency_ms": latency_ms, "message": "模型连接与结构化输出可用"})


async def list_skills(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    with store.session() as session:
        rows = session.exec(
            select(SkillVersion)
            .where(SkillVersion.status == "ENABLED")
            .order_by(SkillVersion.created_at)
        ).all()
    return web.json_response({"items": [_skill_dict(item) for item in rows]})


async def _receive_skill_zip(request: web.Request) -> tuple[Path, dict[str, Any]]:
    settings = request.app[APP_SETTINGS]
    try:
        reader = await request.multipart()
    except (AssertionError, web.HTTPBadRequest) as exc:
        raise WorkbenchError("UPLOAD_MULTIPART_REQUIRED", "请上传 ZIP 格式 Skill 包。", status=400) from exc
    part = await reader.next()
    while part is not None and not part.filename:
        part = await reader.next()
    if part is None or not part.filename or Path(part.filename).suffix.lower() != ".zip":
        raise WorkbenchError("SKILL_ZIP_REQUIRED", "请选择 ZIP 格式 Skill 包。", status=422)
    package_id = new_id()
    temporary = settings.skill_dir / f"{package_id}.uploading"
    final_path = settings.skill_dir / f"{package_id}.zip"
    size = 0
    try:
        async with aiofiles.open(temporary, "wb") as stream:
            while chunk := await part.read_chunk(size=512 * 1024):
                size += len(chunk)
                if size > 20 * 1024 * 1024:
                    raise WorkbenchError("SKILL_ZIP_TOO_LARGE", "Skill ZIP 不能超过 20 MB。", status=413)
                await stream.write(chunk)
        manifest = validate_skill_zip(temporary)
        temporary.replace(final_path)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return final_path, manifest


def _instance_manifest(
    *,
    lineage_id: str,
    source: SkillVersion | None,
    scene_name: str,
    notes: str,
    package_manifest: dict[str, Any] | None = None,
) -> dict[str, Any]:
    source_manifest = _json_loads(source.manifest_json, {}) if source else {}
    return {
        **(package_manifest or {}),
        "built_in": False,
        "kind": "INSTANCE",
        "read_only": False,
        "lineage_id": lineage_id,
        "source_skill_id": source.id if source else None,
        "source_name": source.name if source else str(source_manifest.get("source_name", "")),
        "scene_name": scene_name,
        "notes": notes,
    }


def _next_skill_version(version: str) -> str:
    try:
        major, minor, _patch = (int(part) for part in version.split(".", maxsplit=2))
    except (TypeError, ValueError):
        return "1.1.0"
    return f"{major}.{minor + 1}.0"


async def upload_skill(request: web.Request) -> web.Response:
    final_path, manifest = await _receive_skill_zip(request)
    skill = SkillVersion(
        name=manifest["name"][:120],
        description=manifest["description"],
        version=str(request.query.get("version", "1.0.0"))[:40],
        package_path=str(final_path),
    )
    skill.manifest_json = json.dumps(
        _instance_manifest(
            lineage_id=skill.id,
            source=None,
            scene_name=str(request.query.get("scene_name", "通用场景"))[:160],
            notes="由本地 ZIP 导入",
            package_manifest=manifest,
        ),
        ensure_ascii=False,
    )
    store = request.app[APP_STORE]
    with store.session() as session:
        session.add(skill)
        session.commit()
        session.refresh(skill)
        session.expunge(skill)
    return web.json_response(_skill_dict(skill), status=201)


async def create_skill_instance(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    source_id = str(payload.get("source_skill_id") or payload.get("template_id") or "").strip()
    if not source_id:
        raise WorkbenchError("SKILL_SOURCE_REQUIRED", "请选择一个 Skill 模板或实例。", status=422)
    store = request.app[APP_STORE]
    with store.session() as session:
        source = session.get(SkillVersion, source_id)
        if source is None:
            raise WorkbenchError("SKILL_NOT_FOUND", "Skill 模板或实例不存在。", status=404)
        source_manifest = _json_loads(source.manifest_json, {})
        name = str(payload.get("name", f"{source.name} 实例")).strip()
        if not name:
            raise WorkbenchError("SKILL_NAME_REQUIRED", "请输入 Skill 实例名称。", status=422)
        skill = SkillVersion(
            name=name[:120],
            description=str(payload.get("description", source.description)).strip(),
            version="0.1.0",
            package_path=source.package_path,
        )
        template_source = source
        if source_manifest.get("kind") == "INSTANCE" and source_manifest.get("source_skill_id"):
            template_source = session.get(SkillVersion, str(source_manifest["source_skill_id"])) or source
        skill.manifest_json = json.dumps(
            _instance_manifest(
                lineage_id=skill.id,
                source=template_source,
                scene_name=str(payload.get("scene_name", "通用场景")).strip()[:160] or "通用场景",
                notes=str(payload.get("notes", "由模板复制，可独立维护版本。"))[:500],
            ),
            ensure_ascii=False,
        )
        session.add(skill)
        session.commit()
        session.refresh(skill)
        session.expunge(skill)
    return web.json_response(_skill_dict(skill), status=201)


async def update_skill(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    store = request.app[APP_STORE]
    with store.session() as session:
        skill = session.get(SkillVersion, request.match_info["skill_id"])
        if skill is None:
            raise WorkbenchError("SKILL_NOT_FOUND", "Skill 不存在。", status=404)
        manifest = _json_loads(skill.manifest_json, {})
        if manifest.get("read_only"):
            raise WorkbenchError("SKILL_TEMPLATE_READ_ONLY", "内置 Skill 模板为只读，请先复制为实例。", status=409)
        if "name" in payload:
            name = str(payload["name"]).strip()
            if not name:
                raise WorkbenchError("SKILL_NAME_REQUIRED", "请输入 Skill 实例名称。", status=422)
            skill.name = name[:120]
        if "description" in payload:
            skill.description = str(payload["description"]).strip()
        if "scene_name" in payload:
            manifest["scene_name"] = str(payload["scene_name"]).strip()[:160] or "通用场景"
        if "notes" in payload:
            manifest["notes"] = str(payload["notes"]).strip()[:500]
        skill.manifest_json = json.dumps(manifest, ensure_ascii=False)
        session.add(skill)
        session.commit()
        session.refresh(skill)
        session.expunge(skill)
    return web.json_response(_skill_dict(skill))


async def list_skill_versions(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    with store.session() as session:
        skill = session.get(SkillVersion, request.match_info["skill_id"])
        if skill is None:
            raise WorkbenchError("SKILL_NOT_FOUND", "Skill 不存在。", status=404)
        lineage_id = str(_json_loads(skill.manifest_json, {}).get("lineage_id", skill.id))
        rows = [
            item
            for item in session.exec(select(SkillVersion).order_by(SkillVersion.created_at.desc())).all()
            if str(_json_loads(item.manifest_json, {}).get("lineage_id", item.id)) == lineage_id
        ]
    return web.json_response({"items": [_skill_dict(item) for item in rows]})


async def upload_skill_version(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    current = store.get(SkillVersion, request.match_info["skill_id"], code="SKILL_NOT_FOUND")
    current_manifest = _json_loads(current.manifest_json, {})
    if current_manifest.get("read_only"):
        raise WorkbenchError("SKILL_TEMPLATE_READ_ONLY", "内置 Skill 模板不能上传新版本。", status=409)
    final_path, package_manifest = await _receive_skill_zip(request)
    with store.session() as session:
        current = session.get(SkillVersion, current.id)
        if current is None:
            final_path.unlink(missing_ok=True)
            raise WorkbenchError("SKILL_NOT_FOUND", "Skill 不存在。", status=404)
        current_manifest = _json_loads(current.manifest_json, {})
        version = str(request.query.get("version", _next_skill_version(current.version))).strip()
        new_skill = SkillVersion(
            name=str(package_manifest.get("name") or current.name)[:120],
            description=str(package_manifest.get("description") or current.description),
            version=version[:40],
            package_path=str(final_path),
            manifest_json=json.dumps(
                _instance_manifest(
                    lineage_id=str(current_manifest.get("lineage_id", current.id)),
                    source=session.get(SkillVersion, str(current_manifest.get("source_skill_id")))
                    if current_manifest.get("source_skill_id")
                    else None,
                    scene_name=str(current_manifest.get("scene_name", "通用场景")),
                    notes=str(request.query.get("notes", "上传新版本"))[:500],
                    package_manifest=package_manifest,
                ),
                ensure_ascii=False,
            ),
        )
        current.status = "SUPERSEDED"
        session.add(current)
        session.add(new_skill)
        session.flush()
        for mount in session.exec(select(AbilityMount).where(AbilityMount.skill_version_id == current.id)).all():
            mount.skill_version_id = new_skill.id
            mount.updated_at = utc_now()
            session.add(mount)
        for profile in session.exec(
            select(AbilityProfile).where(AbilityProfile.skill_version_id == current.id)
        ).all():
            profile.skill_version_id = new_skill.id
            profile.updated_at = utc_now()
            session.add(profile)
        session.commit()
        session.refresh(new_skill)
        session.expunge(new_skill)
    return web.json_response(_skill_dict(new_skill), status=201)


def _generated_skill_zip(skill: SkillVersion) -> bytes:
    manifest = _json_loads(skill.manifest_json, {})
    skill_slug = str(manifest.get("slug") or skill.name).strip().lower().replace(" ", "-")
    skill_md = (
        "---\n"
        f"name: {skill_slug or 'knowledge-workbench-skill'}\n"
        f"description: {skill.description or skill.name}\n"
        "---\n\n"
        f"# {skill.name}\n\n"
        "该包由知识萃取智能体工作台生成，包含可继续维护的 Skill 骨架。\n"
    )
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("SKILL.md", skill_md)
        archive.writestr("references/README.md", "# References\n\n在此补充已审定知识和来源说明。\n")
        archive.writestr("scripts/validate.py", "print('skill package ready')\n")
        archive.writestr("assets/schema.json", json.dumps({"version": skill.version}, ensure_ascii=False))
    return buffer.getvalue()


async def download_skill(request: web.Request) -> web.StreamResponse:
    skill = request.app[APP_STORE].get(
        SkillVersion,
        request.match_info["skill_id"],
        code="SKILL_NOT_FOUND",
    )
    headers = {"Content-Disposition": f'attachment; filename="skill-{skill.id[:8]}-{skill.version}.zip"'}
    path = Path(skill.package_path) if skill.package_path else None
    if path and path.is_file():
        response = web.FileResponse(path, headers=headers)
        response.content_type = "application/zip"
        return response
    return web.Response(body=_generated_skill_zip(skill), content_type="application/zip", headers=headers)


def _ability_items(session: Any, scope_key: str) -> list[dict[str, Any]]:
    mounts = {item.ability_key: item for item in session.exec(select(AbilityMount)).all()}
    profiles = {
        item.mount_id: item
        for item in session.exec(select(AbilityProfile).where(AbilityProfile.scope_key == scope_key)).all()
    }
    models = {
        item.id: _model_dict(item)
        for item in session.exec(select(ModelConnection).where(ModelConnection.provider != "FakeModel")).all()
    }
    skills = {item.id: _skill_dict(item) for item in session.exec(select(SkillVersion)).all()}
    items = []
    for spec in ABILITY_SPECS:
        mount = mounts.get(spec["key"])
        if mount is None:
            continue
        profile = profiles.get(mount.id) if scope_key != "GLOBAL" else None
        enabled = profile.enabled if profile else mount.enabled
        model_id = profile.model_connection_id if profile else mount.model_connection_id
        skill_id = profile.skill_version_id if profile else mount.skill_version_id
        params_json = profile.params_json if profile else mount.params_json
        items.append(
            {
                "id": mount.id,
                "profile_id": profile.id if profile else None,
                "scope_key": scope_key,
                "inherited": scope_key != "GLOBAL" and profile is None,
                "ability_key": mount.ability_key,
                "display_name": mount.display_name,
                "description": mount.description,
                "stage": spec["stage"],
                "trigger": spec["trigger"],
                "location": spec["location"],
                "enabled": enabled,
                "model_connection_id": model_id,
                "skill_version_id": skill_id,
                "model": models.get(str(model_id)),
                "skill": skills.get(str(skill_id)),
                "params": _json_loads(params_json, {}),
                "updated_at": _iso(profile.updated_at if profile else mount.updated_at),
            }
        )
    return items


def _apply_ability_payload(
    session: Any,
    mount: AbilityMount,
    scope_key: str,
    payload: dict[str, Any],
) -> None:
    profile = None
    if scope_key != "GLOBAL":
        profile = session.exec(
            select(AbilityProfile).where(
                AbilityProfile.mount_id == mount.id,
                AbilityProfile.scope_key == scope_key,
            )
        ).first()
        if profile is None:
            profile = AbilityProfile(
                mount_id=mount.id,
                scope_key=scope_key,
                enabled=mount.enabled,
                model_connection_id=mount.model_connection_id,
                skill_version_id=mount.skill_version_id,
                params_json=mount.params_json,
            )

    target = profile or mount
    if "model_connection_id" in payload:
        model_id = payload["model_connection_id"] or None
        model = session.get(ModelConnection, model_id) if model_id else None
        if model_id and (
            model is None
            or model.provider == "FakeModel"
            or not model.enabled
            or not model.encrypted_api_key
        ):
            raise WorkbenchError("MODEL_NOT_FOUND", "模型连接不存在或不可用。", status=404)
        target.model_connection_id = model_id
    if "skill_version_id" in payload:
        skill_id = payload["skill_version_id"] or None
        skill = session.get(SkillVersion, skill_id) if skill_id else None
        if skill_id and (skill is None or skill.status != "ENABLED"):
            raise WorkbenchError("SKILL_NOT_FOUND", "Skill 不存在或已停用。", status=404)
        target.skill_version_id = skill_id
    if "enabled" in payload:
        target.enabled = bool(payload["enabled"])
    if "params" in payload:
        params = payload["params"]
        if not isinstance(params, dict):
            raise WorkbenchError("ABILITY_PARAMS_INVALID", "能力参数必须是 JSON 对象。", status=422)
        existing = _json_loads(target.params_json, {})
        target.params_json = json.dumps(existing | params, ensure_ascii=False)
    target.updated_at = utc_now()
    session.add(target)


async def list_ability_scopes(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    items = [{"key": "GLOBAL", "label": "通用场景（默认配置）", "kind": "GLOBAL"}]
    with store.session() as session:
        scenes = session.exec(
            select(Scene).where(Scene.archived_at.is_(None)).order_by(Scene.created_at.desc())
        ).all()
        for scene in scenes:
            items.append({"key": f"SCENE:{scene.id}", "label": scene.name, "kind": "SCENE"})
            latest = session.exec(
                select(ExtractionRound)
                .where(ExtractionRound.scene_id == scene.id)
                .order_by(ExtractionRound.version.desc())
            ).first()
            subscenes = _json_loads(latest.subscenes_json, []) if latest else []
            for index, subscene in enumerate(subscenes):
                if str(subscene).strip():
                    items.append(
                        {
                            "key": f"SUBSCENE:{scene.id}:{index}",
                            "label": f"{scene.name} › {str(subscene).strip()}",
                            "kind": "SUBSCENE",
                        }
                    )
    return web.json_response({"items": items})


async def list_ability_mounts(request: web.Request) -> web.Response:
    store = request.app[APP_STORE]
    scope_key = request.query.get("scope_key", "GLOBAL").strip() or "GLOBAL"
    with store.session() as session:
        items = _ability_items(session, scope_key)
    return web.json_response({"scope_key": scope_key, "items": items})


async def update_ability_mount(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    store = request.app[APP_STORE]
    mount_id = request.match_info["mount_id"]
    scope_key = str(payload.get("scope_key", request.query.get("scope_key", "GLOBAL"))).strip() or "GLOBAL"
    with store.session() as session:
        mount = session.get(AbilityMount, mount_id)
        if mount is None:
            raise WorkbenchError("ABILITY_MOUNT_NOT_FOUND", "能力配置不存在。", status=404)
        _apply_ability_payload(session, mount, scope_key, payload)
        session.commit()
        item = next(row for row in _ability_items(session, scope_key) if row["id"] == mount_id)
    return web.json_response(item)


async def import_ability_configuration(request: web.Request) -> web.Response:
    payload = await _json_body(request)
    scope_key = str(payload.get("scope_key", "GLOBAL")).strip() or "GLOBAL"
    rows = payload.get("items")
    if not isinstance(rows, list):
        raise WorkbenchError("ABILITY_CONFIG_INVALID", "导入配置缺少 items 数组。", status=422)
    store = request.app[APP_STORE]
    with store.session() as session:
        mounts = {item.ability_key: item for item in session.exec(select(AbilityMount)).all()}
        for row in rows:
            if not isinstance(row, dict) or str(row.get("ability_key", "")) not in mounts:
                raise WorkbenchError("ABILITY_CONFIG_INVALID", "导入配置包含未知智能体。", status=422)
            _apply_ability_payload(session, mounts[str(row["ability_key"])], scope_key, row)
        session.commit()
        items = _ability_items(session, scope_key)
    return web.json_response({"scope_key": scope_key, "items": items})


async def reset_ability_mounts(request: web.Request) -> web.Response:
    payload = await _json_body(request) if request.can_read_body and request.content_length else {}
    scope_key = str(payload.get("scope_key", "GLOBAL")).strip() or "GLOBAL"
    requested_model_id = str(payload.get("model_connection_id", "")).strip()
    store = request.app[APP_STORE]
    with store.session() as session:
        models = session.exec(
            select(ModelConnection)
            .where(ModelConnection.provider != "FakeModel", ModelConnection.enabled == True)  # noqa: E712
            .order_by(ModelConnection.created_at)
        ).all()
        available_models = [item for item in models if item.encrypted_api_key]
        if requested_model_id:
            selected_model = next((item for item in available_models if item.id == requested_model_id), None)
            if selected_model is None:
                raise WorkbenchError(
                    "MODEL_NOT_AVAILABLE",
                    "所选模型连接不存在、已停用或缺少 API Key。",
                    status=422,
                )
        elif len(available_models) == 1:
            selected_model = available_models[0]
        elif len(available_models) > 1:
            raise WorkbenchError(
                "MODEL_SELECTION_REQUIRED",
                "存在多个可用模型，请先选择要批量应用的模型。",
                status=422,
                details={"available_count": len(available_models)},
            )
        else:
            selected_model = None
        if selected_model is None:
            raise WorkbenchError("MODEL_REQUIRED", "请先接入并测试一个真实模型。", status=409)
        skills = {}
        for item in session.exec(select(SkillVersion).where(SkillVersion.status == "ENABLED")).all():
            slug = _json_loads(item.manifest_json, {}).get("slug")
            if slug:
                skills[str(slug)] = item
        mounts = {item.ability_key: item for item in session.exec(select(AbilityMount)).all()}
        for spec in ABILITY_SPECS:
            mount = mounts[spec["key"]]
            skill = skills.get(spec["skill_slug"])
            _apply_ability_payload(
                session,
                mount,
                scope_key,
                {
                    "enabled": True,
                    "model_connection_id": selected_model.id,
                    "skill_version_id": skill.id if skill else None,
                    "params": spec["defaults"],
                },
            )
        session.commit()
        items = _ability_items(session, scope_key)
    return web.json_response({"reset": True, "scope_key": scope_key, "items": items})


async def serve_index(request: web.Request) -> web.StreamResponse:
    dist = request.app[APP_SETTINGS].frontend_dist
    index = dist / "index.html"
    if not index.is_file():
        return web.json_response(
            {
                "message": "Frontend is not built yet.",
                "hint": "Run `npm install && npm run build` in examples/knowledge_extraction_workbench/frontend.",
            },
            status=503,
        )
    return web.FileResponse(index)


async def serve_static(request: web.Request) -> web.StreamResponse:
    dist = request.app[APP_SETTINGS].frontend_dist.resolve()
    relative = request.match_info.get("path", "")
    candidate = (dist / relative).resolve()
    if candidate != dist and dist not in candidate.parents:
        raise WorkbenchError("STATIC_PATH_INVALID", "静态资源路径无效。", status=404)
    if candidate.is_file():
        return web.FileResponse(candidate)
    return await serve_index(request)


def create_app(settings: Settings | None = None, *, test_model: Any | None = None) -> web.Application:
    settings = settings or Settings.from_env()
    settings.ensure_directories()
    store = Store(settings.database_path)
    store.initialize()
    secret_box = SecretBox(settings.key_path)
    service = WorkbenchService(settings, store, secret_box, test_model=test_model)
    app = web.Application(middlewares=[error_middleware], client_max_size=settings.max_upload_bytes + 1024 * 1024)
    app[APP_SETTINGS] = settings
    app[APP_STORE] = store
    app[APP_SERVICE] = service
    app[APP_SECRET_BOX] = secret_box

    app.router.add_get("/api/v1/health", health)
    app.router.add_get("/api/v1/dashboard", dashboard)
    app.router.add_get("/api/v1/scenes", list_scenes)
    app.router.add_post("/api/v1/scenes", create_scene)
    app.router.add_get("/api/v1/scenes/{scene_id}", get_scene)
    app.router.add_patch("/api/v1/scenes/{scene_id}", update_scene)
    app.router.add_delete("/api/v1/scenes/{scene_id}", archive_scene)
    app.router.add_get("/api/v1/scenes/{scene_id}/rounds", list_rounds)
    app.router.add_post("/api/v1/scenes/{scene_id}/rounds", create_round)
    app.router.add_get("/api/v1/rounds/{round_id}/materials", list_materials)
    app.router.add_post("/api/v1/rounds/{round_id}/materials", upload_round_material)
    app.router.add_patch("/api/v1/materials/{material_id}", update_material)
    app.router.add_get("/api/v1/explorations", list_explorations)
    app.router.add_post("/api/v1/explorations", create_exploration)
    app.router.add_delete("/api/v1/explorations/{exploration_id}", archive_exploration)
    app.router.add_get("/api/v1/explorations/{exploration_id}/materials", list_exploration_materials)
    app.router.add_post("/api/v1/explorations/{exploration_id}/materials", upload_exploration_material)
    app.router.add_post("/api/v1/explorations/{exploration_id}/analyze", analyze_exploration)
    app.router.add_get("/api/v1/explorations/{exploration_id}/candidates", list_candidates)
    app.router.add_post(
        "/api/v1/explorations/{exploration_id}/candidates/{candidate_id}/create-scene", candidate_create_scene
    )
    app.router.add_post("/api/v1/rounds/{round_id}/extract", start_extraction)
    app.router.add_get("/api/v1/jobs/{job_id}", get_job)
    app.router.add_get("/api/v1/jobs/{job_id}/events", job_events)
    app.router.add_post("/api/v1/jobs/{job_id}/retry", retry_job)
    app.router.add_get("/api/v1/rounds/{round_id}/document", get_document)
    app.router.add_put("/api/v1/rounds/{round_id}/document", save_document)
    app.router.add_get("/api/v1/rounds/{round_id}/revisions", list_revisions)
    app.router.add_get("/api/v1/rounds/{round_id}/suggestions", list_suggestions)
    app.router.add_post("/api/v1/rounds/{round_id}/suggestions", create_suggestion)
    app.router.add_post("/api/v1/suggestions/{suggestion_id}/apply", apply_suggestion)
    app.router.add_post("/api/v1/suggestions/{suggestion_id}/reject", reject_suggestion)
    app.router.add_get("/api/v1/rounds/{round_id}/assets", list_assets)
    app.router.add_post("/api/v1/rounds/{round_id}/assets", generate_round_assets)
    app.router.add_get("/api/v1/rounds/{round_id}/assets/download", download_round_assets)
    app.router.add_get("/api/v1/assets/{asset_id}/download", download_asset)
    app.router.add_get("/api/v1/assets/{asset_id}/preview", preview_asset)
    app.router.add_post("/api/v1/rounds/{round_id}/publish", publish_round)
    app.router.add_get("/api/v1/runtime/skills", list_runtime_skills)
    app.router.add_post("/api/v1/runtime/tryouts", runtime_tryout)
    app.router.add_post("/api/v1/runtime/tryouts/upload", runtime_tryout_upload)
    app.router.add_get("/api/v1/evaluations", list_evaluations)
    app.router.add_post("/api/v1/evaluations", create_evaluation)
    app.router.add_post("/api/v1/evaluations/upload", upload_evaluation)
    app.router.add_get("/api/v1/evaluations/{evaluation_id}", get_evaluation)
    app.router.add_post("/api/v1/evaluations/{evaluation_id}/feedback", evaluation_to_feedback)
    app.router.add_get("/api/v1/feedback-tasks", list_feedback_tasks)
    app.router.add_post("/api/v1/feedback-tasks", create_feedback_task)
    app.router.add_get("/api/v1/feedback-tasks/{task_id}", get_feedback_task)
    app.router.add_put("/api/v1/feedback-tasks/{task_id}", save_feedback_task)
    app.router.add_post("/api/v1/feedback-tasks/{task_id}/cases", upload_feedback_cases)
    app.router.add_post("/api/v1/feedback-tasks/{task_id}/analyze", analyze_feedback_task)
    app.router.add_get("/api/v1/feedback-tasks/{task_id}/export", export_feedback_task)
    app.router.add_post("/api/v1/feedback-tasks/{task_id}/promote", promote_feedback_task)
    app.router.add_get("/api/v1/models", list_models)
    app.router.add_post("/api/v1/models", create_model)
    app.router.add_put("/api/v1/models/{model_id}", update_model)
    app.router.add_delete("/api/v1/models/{model_id}", delete_model)
    app.router.add_post("/api/v1/models/{model_id}/test", test_model_connection)
    app.router.add_get("/api/v1/skills", list_skills)
    app.router.add_post("/api/v1/skills", upload_skill)
    app.router.add_post("/api/v1/skills/instances", create_skill_instance)
    app.router.add_put("/api/v1/skills/{skill_id}", update_skill)
    app.router.add_get("/api/v1/skills/{skill_id}/versions", list_skill_versions)
    app.router.add_post("/api/v1/skills/{skill_id}/versions", upload_skill_version)
    app.router.add_get("/api/v1/skills/{skill_id}/download", download_skill)
    app.router.add_get("/api/v1/ability-scopes", list_ability_scopes)
    app.router.add_get("/api/v1/ability-mounts", list_ability_mounts)
    app.router.add_put("/api/v1/ability-mounts/configuration", import_ability_configuration)
    app.router.add_post("/api/v1/ability-mounts/defaults", reset_ability_mounts)
    app.router.add_put("/api/v1/ability-mounts/{mount_id}", update_ability_mount)
    app.router.add_route("*", "/api/{path:.*}", api_not_found)
    app.router.add_get("/", serve_index)
    app.router.add_get("/{path:.*}", serve_static)

    async def _cleanup(_: web.Application) -> None:
        await service.close()

    app.on_cleanup.append(_cleanup)
    return app
